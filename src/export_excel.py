"""Export the final Excel deliverable: influencers + segment + messages + summary."""
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import ENRICHED_CSV, MESSAGES_CSV, SEGMENT_CSV

OUT = "output/indian_micro_influencers.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def read(path):
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def write_table(ws, rows, cols, widths):
    for c, (header, _) in enumerate(cols, 1):
        cell = ws.cell(1, c, header)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    for r, row in enumerate(rows, 2):
        for c, (_, key) in enumerate(cols, 1):
            val = row[key]
            if key == "followers":
                val = int(val)
            elif key == "engagement_rate_pct":
                val = float(val) / 100  # store as fraction, render as %
            elif key == "brand_fit_score":
                val = int(val)
            cell = ws.cell(r, c, val)
            cell.font, cell.border = BODY_FONT, THIN
            cell.alignment = Alignment(vertical="top", wrap_text=key in
                                       ("content_themes", "email_body", "instagram_dm", "email_source"))
            if key == "followers":
                cell.number_format = "#,##0"
            elif key == "engagement_rate_pct":
                cell.number_format = "0.00%"
    ws.freeze_panes = "A2"


INF_COLS = [
    ("Name", "name"), ("Handle", "handle"), ("Platform(s)", "platforms"),
    ("Followers", "followers"), ("Engagement Rate", "engagement_rate_pct"),
    ("ER Source", "er_source"), ("Niche", "niche"), ("Content Themes", "content_themes"),
    ("Contact Email", "contact_email"), ("Email Status", "email_status"),
    ("Email Source", "email_source"), ("Profile URL", "profile_url"),
    ("City/State", "city_state"), ("Audience Indicator", "audience_indicator"),
    ("Brand Fit (0-100)", "brand_fit_score"), ("Discovered Via", "discovered_via"),
]
INF_W = [24, 22, 20, 11, 11, 24, 10, 32, 32, 16, 30, 40, 22, 30, 12, 30]

MSG_COLS = [
    ("Name", "name"), ("Handle", "handle"), ("Niche", "niche"),
    ("Contact Email", "contact_email"), ("Email Subject", "subject"),
    ("Email Body (60-90 words)", "email_body"), ("Email Words", "email_words"),
    ("Instagram DM (15-30 words)", "instagram_dm"), ("DM Words", "dm_words"),
]
MSG_W = [24, 22, 10, 30, 42, 80, 9, 60, 8]


def main():
    infl, seg, msgs = read(ENRICHED_CSV), read(SEGMENT_CSV), read(MESSAGES_CSV)
    wb = Workbook()

    ws = wb.active
    ws.title = "All Influencers"
    write_table(ws, infl, INF_COLS, INF_W)

    ws2 = wb.create_sheet("Segment Lifestyle+Travel")
    write_table(ws2, seg, INF_COLS, INF_W)

    ws3 = wb.create_sheet("Outreach Messages")
    write_table(ws3, msgs, MSG_COLS, MSG_W)

    ws4 = wb.create_sheet("Summary")
    n = len(infl)
    labels = [
        ("Total influencers", f"=COUNTA('All Influencers'!A2:A{n + 1})"),
        ("With verified public email", f"=COUNTIF('All Influencers'!J2:J{n + 1},\"verified_public\")"),
        ("Lifestyle", f"=COUNTIF('All Influencers'!G2:G{n + 1},\"Lifestyle\")"),
        ("Travel", f"=COUNTIF('All Influencers'!G2:G{n + 1},\"Travel\")"),
        ("Tech", f"=COUNTIF('All Influencers'!G2:G{n + 1},\"Tech\")"),
        ("Food", f"=COUNTIF('All Influencers'!G2:G{n + 1},\"Food\")"),
        ("Fitness", f"=COUNTIF('All Influencers'!G2:G{n + 1},\"Fitness\")"),
        ("Average followers", f"=ROUND(AVERAGE('All Influencers'!D2:D{n + 1}),0)"),
        ("Average engagement rate", f"=AVERAGE('All Influencers'!E2:E{n + 1})"),
        ("Segment size (Lifestyle+Travel, email, ER>=2%)",
         f"=COUNTA('Segment Lifestyle+Travel'!A2:A{len(seg) + 1})"),
    ]
    ws4.cell(1, 1, "Campaign Summary").font = Font(name="Arial", bold=True, size=13)
    for i, (label, formula) in enumerate(labels, 3):
        ws4.cell(i, 1, label).font = BODY_FONT
        c = ws4.cell(i, 2, formula)
        c.font = Font(name="Arial", bold=True, size=10)
        if "engagement" in label:
            c.number_format = "0.00%"
        elif "followers" in label.lower():
            c.number_format = "#,##0"
    ws4.column_dimensions["A"].width = 45
    ws4.column_dimensions["B"].width = 16
    note = ("Note: follower counts & engagement rates as reported by public directories "
            "(Feedspot/StarNgage), July 2026. Missing ERs use India micro-tier benchmarks, "
            "labelled 'estimated' in ER Source. Emails only where publicly listed by the creator.")
    ws4.cell(len(labels) + 5, 1, note).font = Font(name="Arial", italic=True, size=9)
    ws4.cell(len(labels) + 5, 1).alignment = Alignment(wrap_text=True)

    wb.save(OUT)
    print(f"-> {OUT}")


if __name__ == "__main__":
    main()
